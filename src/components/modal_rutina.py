# python
import io
import openpyxl
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment as XlAlignment, Border, Side
import flet as ft
from flet import FilePicker, FilePickerFileType
from constants import *
from database_orm import Cronograma

DIAS_CORTOS = ["Lun", "mar", "mie", "jue", "vie", "sab", "dom"]
DIAS_COMPLETOS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def crear_modal_rutina(nombre_miembro: str, page: ft.Page, session, miembro_id: int, on_guardar=None):
    is_mobile = page.width is not None and page.width < 600

    # ── Limpiar rutinas viejas sin semana ──────────────────────
    session.query(Cronograma).filter_by(miembro_id=miembro_id, semana=None).delete()
    session.commit()

    # ── Estado mutable ──────────────────────────────────────────
    state = {
        "semana_guardada": [set(), set(), set(), set()],
        "semana_pending": [set(), set(), set(), set()],
    }

    # ── Helpers de BD ───────────────────────────────────────────
    def _cargar_rutinas():
        return session.query(Cronograma).filter_by(miembro_id=miembro_id).all()

    for r in _cargar_rutinas():
        if r.repetir_semanal and r.semana is not None:
            state["semana_guardada"][r.semana - 1].add(r.dia_semana)

    # ── Header ───────────────────────────────────────────────────
    header = ft.Container(
        content=ft.Text(
            f"Rutina de {nombre_miembro}",
            size=14, weight="bold",
            color=THEME_TEXT_PRIMARY,
            text_align=ft.TextAlign.CENTER,
        ),
        bgcolor=THEME_HEADER_BG_MODAL,
        padding=ft.Padding(20, 14, 20, 14),
        alignment=ft.Alignment.CENTER,
        border_radius=ft.BorderRadius.only(top_left=12, top_right=12),
    )

    # ── Contenedores mutables ───────────────────────────────────
    dias_container = ft.Container(expand=True)
    rutinas_lista = ft.Container()

    # ── Construir selector de 4 semanas ────────────────────────
    def _construir_dias():
        labels_semana = ["Semana 1", "Semana 2", "Semana 3", "Semana 4"]

        semana_blocks = []
        for idx, label in enumerate(labels_semana):
            guardados = state["semana_guardada"][idx]
            pending = state["semana_pending"][idx]
            buttons = []
            for i, nombre in enumerate(DIAS_CORTOS):
                es_guardado = i in guardados
                es_pending = i in pending

                if es_guardado or es_pending:
                    bg = BTNS_BG
                    txt_color = THEME_TEAL_TEXT
                else:
                    bg = "#3A3D42"
                    txt_color = THEME_TEXT_PRIMARY

                click_fn = None if es_guardado else lambda e, s=idx, wd=i: _toggle_pending(s, wd)

                buttons.append(
                    ft.Container(
                        content=ft.Text(
                            nombre,
                            color=txt_color,
                            size=11,
                            weight="bold",
                            text_align=ft.TextAlign.CENTER,
                        ),
                        expand=True,
                        height=36,
                        border_radius=6,
                        bgcolor=bg,
                        alignment=ft.Alignment.CENTER,
                        on_click=click_fn,
                    )
                )

            semana_blocks.append(
                ft.Column(
                    controls=[
                        ft.Text(label, size=12, weight="bold", color=THEME_TEXT_PRIMARY),
                        ft.Row(controls=buttons, spacing=4),
                    ],
                    spacing=4,
                )
            )

        dias_container.content = ft.Column(
            controls=semana_blocks,
            spacing=10,
        )

    def _toggle_pending(semana, wd):
        if wd in state["semana_pending"][semana]:
            state["semana_pending"][semana].remove(wd)
        else:
            state["semana_pending"][semana].add(wd)
        _construir_dias()
        page.update()

    # ── Lista de rutinas existentes ──────────────────────────────
    def _eliminar_rutina(rutina_id):
        r = session.query(Cronograma).filter_by(id=rutina_id).first()
        if r:
            session.delete(r)
            session.commit()
        _actualizar_rutinas()
        page.update()

    def _construir_lista_rutinas():
        rutinas = _cargar_rutinas()
        if not rutinas:
            return ft.Text(
                "Sin rutinas registradas",
                size=13, color=THEME_TEXT_SECONDARY, italic=True,
            )

        items = []
        for r in rutinas:
            dia_nombre = DIAS_COMPLETOS[r.dia_semana]
            semana_str = f" (Semana {r.semana})" if r.semana else ""
            desc = r.descripcion or "—"
            items.append(
                ft.Container(
                    content=ft.Row(
                        controls=[
                            ft.Text(
                                f"{dia_nombre}{semana_str}  ·  {desc}",
                                color=THEME_TEXT_PRIMARY,
                                size=13,
                                expand=True,
                            ),
                            ft.IconButton(
                                icon=ft.Icons.CLOSE,
                                icon_color=ft.Colors.RED_400,
                                icon_size=16,
                                tooltip="Eliminar rutina",
                                on_click=lambda e, rid=r.id: _eliminar_rutina(rid),
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    ),
                    bgcolor="#3A3D42",
                    border_radius=8,
                    padding=ft.Padding(10, 8, 4, 8),
                    margin=ft.Margin.only(bottom=4),
                )
            )
        return ft.Column(controls=items, spacing=0, scroll=ft.ScrollMode.AUTO)

    def _actualizar_rutinas():
        state["semana_guardada"] = [set(), set(), set(), set()]
        for r in _cargar_rutinas():
            if r.repetir_semanal and r.semana is not None:
                state["semana_guardada"][r.semana - 1].add(r.dia_semana)
        rutinas_lista.content = _construir_lista_rutinas()
        _construir_dias()

    # ── Descripción ──────────────────────────────────────────────
    descripcion_field = ft.TextField(
        label="Descripción",
        border_color=THEME_BORDER_COLOR,
        focused_border_color=THEME_TEAL,
        color=THEME_TEXT_PRIMARY,
        label_style=ft.TextStyle(color=THEME_TEXT_SECONDARY),
        cursor_color=THEME_TEAL,
        max_length=200,
        multiline=True,
        min_lines=1,
        max_lines=3,
        expand=True,
    )

    # ── Cerrar ───────────────────────────────────────────────────
    def cerrar_modal(e=None):
        modal_rutina.open = False
        page.update()
        if on_guardar:
            on_guardar()

    # ── Agregar ──────────────────────────────────────────────────
    def _agregar(e):
        descripcion = (descripcion_field.value or "").strip()
        if not descripcion:
            page.snack_bar = ft.SnackBar(
                content=ft.Text("Escribe una descripción primero"), duration=2500
            )
            page.snack_bar.open = True
            page.update()
            return

        for semana in range(4):
            for wd in state["semana_pending"][semana]:
                existente = session.query(Cronograma).filter_by(
                    miembro_id=miembro_id, dia_semana=wd, semana=semana + 1, repetir_semanal=True,
                ).first()
                if existente:
                    existente.descripcion = descripcion
                else:
                    session.add(Cronograma(
                        miembro_id=miembro_id, dia_semana=wd, semana=semana + 1,
                        descripcion=descripcion, repetir_semanal=True,
                    ))
                state["semana_guardada"][semana].add(wd)
        session.commit()

        state["semana_pending"] = [set(), set(), set(), set()]
        descripcion_field.value = ""
        rutinas_lista.content = _construir_lista_rutinas()
        _construir_dias()
        page.update()

    # ── Generar Excel ──────────────────────────────────────────
    async def _generar_excel(e):
        if not any(state["semana_guardada"]) and not any(state["semana_pending"]):
            page.snack_bar = ft.SnackBar(
                content=ft.Text("No hay rutinas para exportar"), duration=2500
            )
            page.snack_bar.open = True
            page.update()
            return

        rutinas = _cargar_rutinas()
        rutinas_dict = {}
        for r in rutinas:
            if r.repetir_semanal and r.semana is not None:
                rutinas_dict[(r.semana - 1, r.dia_semana)] = r.descripcion or ""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rutina"

        dias = ["Semana", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

        header_fill = PatternFill(start_color="1B6CA8", end_color="1B6CA8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        center_align = XlAlignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, name in enumerate(dias, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        colores_semana = [
            "7ED7F1",
            "A8E663",
            "FFB347",
            "FFE066",
        ]

        cuerpo_font = Font(color="000000", size=11)

        for semana in range(4):
            fill = PatternFill(start_color=colores_semana[semana], end_color=colores_semana[semana], fill_type="solid")
            row_num = semana + 2
            ws.cell(row=row_num, column=1, value=f"Semana {semana + 1}").fill = fill
            ws.cell(row=row_num, column=1).font = cuerpo_font
            ws.cell(row=row_num, column=1).alignment = center_align
            ws.cell(row=row_num, column=1).border = thin_border
            for dia_idx in range(7):
                if dia_idx in state["semana_guardada"][semana]:
                    desc = rutinas_dict.get((semana, dia_idx), "")
                elif dia_idx in state["semana_pending"][semana]:
                    desc = (descripcion_field.value or "").strip()
                else:
                    desc = ""
                cell = ws.cell(row=row_num, column=dia_idx + 2, value=desc)
                cell.fill = fill
                cell.font = cuerpo_font
                cell.alignment = center_align
                cell.border = thin_border

        ws.column_dimensions['A'].width = 12
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()

        name_clean = nombre_miembro.replace(' ', '_')
        name_clean = "".join(c for c in name_clean if c.isalnum() or c in ('_', '-'))
        fecha_str = datetime.now().strftime("%Y%m%d")
        nombre_archivo = f"Rutina_{name_clean}_{fecha_str}.xlsx"

        picker = FilePicker()
        es_mobile = page.platform is not None and page.platform.is_mobile()

        if page.web or es_mobile:
            ruta = await picker.save_file(
                dialog_title="Guardar Excel de Rutina",
                file_name=nombre_archivo,
                file_type=FilePickerFileType.CUSTOM,
                allowed_extensions=["xlsx"],
                src_bytes=excel_bytes,
            )
        else:
            ruta = await picker.save_file(
                dialog_title="Guardar Excel de Rutina",
                file_name=nombre_archivo,
                file_type=FilePickerFileType.CUSTOM,
                allowed_extensions=["xlsx"],
            )
            if ruta:
                Path(ruta).write_bytes(excel_bytes)

        if not ruta:
            return

        page.snack_bar = ft.SnackBar(
            content=ft.Text("Excel creado!"),
            bgcolor=THEME_TEAL,
            duration=3000,
        )
        page.snack_bar.open = True
        page.update()

    salir_btn = ft.OutlinedButton(
        content=ft.Text("Salir"),
        style=ft.ButtonStyle(
            color=THEME_TEXT_PRIMARY,
            side=ft.BorderSide(1, THEME_BORDER_COLOR),
            mouse_cursor=ft.MouseCursor.CLICK,
        ),
        on_click=lambda e: cerrar_modal(),
    )

    agregar_btn = ft.FilledButton(
        content=ft.Text("Agregar"),
        bgcolor=BTNS_BG,
        color=THEME_TEAL_TEXT,
        style=ft.ButtonStyle(mouse_cursor=ft.MouseCursor.CLICK),
        on_click=_agregar,
    )

    excel_btn = ft.FilledButton(
        content=ft.Row(
            controls=[ft.Icon(ft.Icons.TABLE_VIEW, size=16, color=THEME_TEAL_TEXT), ft.Text("Excel")],
            spacing=4,
            alignment=ft.MainAxisAlignment.CENTER,
        ),
        bgcolor=BTNS_BG,
        color=THEME_TEAL_TEXT,
        style=ft.ButtonStyle(mouse_cursor=ft.MouseCursor.CLICK),
        on_click=_generar_excel,
    )

    # ── Construir bloques ────────────────────────────────────────
    _construir_dias()
    rutinas_lista.content = _construir_lista_rutinas()

    bloque_calendario = ft.Container(
        content=dias_container,
        padding=ft.Padding(12, 8, 12, 8),
        bgcolor=THEME_CARD_BG,
        border_radius=12,
    )

    bloque_central = ft.Container(
        content=ft.Column(
            controls=[
                ft.Text(
                    "Rutinas actuales:", size=13, weight="bold",
                    color=THEME_TEXT_PRIMARY,
                ),
                rutinas_lista,
                ft.Container(
                    content=descripcion_field,
                    alignment=ft.Alignment.CENTER,
                    expand=True,
                ),
            ],
            spacing=8,
        ),
    )

    bloque_botones = ft.Container(
        content=ft.Row(
            controls=[excel_btn, agregar_btn, salir_btn],
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=10,
        ),
        padding=ft.Padding.only(top=4),
    )

    contenido_principal = ft.Column(
        controls=[
            header,
            ft.Container(
                content=ft.Column(
                    controls=[bloque_calendario, bloque_central, bloque_botones],
                    spacing=12,
                ),
                padding=ft.Padding(16, 12, 16, 16),
            ),
        ],
        spacing=0,
        scroll=ft.ScrollMode.AUTO,
    )

    ancho_modal = page.width * 0.95 if is_mobile else 800

    modal_rutina = ft.AlertDialog(
        modal=True,
        bgcolor=THEME_CARD_BG,
        shape=ft.RoundedRectangleBorder(radius=12),
        content=ft.Container(
            content=contenido_principal,
            width=ancho_modal,
        ),
        content_padding=ft.Padding.all(0),
    )

    return modal_rutina
